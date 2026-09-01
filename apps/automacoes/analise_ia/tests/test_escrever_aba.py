"""
=== ARQUIVO: apps/automacoes/analise_ia/tests/test_escrever_aba.py ===
Propósito: Prova que `ggci.escrever_aba` grava exatamente as mesmas células que o
`DataFrame.to_excel` que ela substituiu.
Autor: N/A
Dependências Principais: unittest, pandas, xlsxwriter, openpyxl

POR QUÊ EXISTE: `escrever_aba` existe só por performance — escreve por coluna em vez de
por célula, e nas abas grandes do relatório isso vale 3,5x. Performance não justifica
mudar uma vírgula do artefato final, então a equivalência precisa ser verificada por
teste, não por inspeção.

O QUE COBRE:
  1. Todos os dtypes que aparecem nas abas (Int64 com pd.NA, float com NaN e inf,
     object misto, datetime, bool, texto com acento, string que parece fórmula).
  2. Os dados REAIS do baseline em tests/baseline/relatorio_geral.xlsx, quando ele
     existe — inclusive passando pela `aplicar_formatacao_visual` de verdade.
"""
import os
import unittest

import numpy as np
import pandas as pd
import openpyxl
import xlsxwriter

from apps.automacoes.analise_ia.services import ggci

BASELINE = os.path.join(os.path.dirname(__file__), 'baseline', 'relatorio_geral.xlsx')
LINHAS_AMOSTRA = 3000


def celulas(caminho, aba=None):
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb[aba] if aba else wb.active
    dados = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return dados


def comparar_celulas(caso, a, b):
    """Devolve a lista de divergências legíveis entre duas matrizes de células."""
    problemas = []
    if len(a) != len(b):
        problemas.append(f"{caso}: {len(a)} linhas no to_excel, {len(b)} no escrever_aba")
    for i, (la, lb) in enumerate(zip(a, b)):
        if len(la) != len(lb):
            problemas.append(f"{caso}: linha {i} com {len(la)} colunas vs {len(lb)}")
            continue
        for j, (va, vb) in enumerate(zip(la, lb)):
            if va == vb:
                continue
            if va is None and vb is None:
                continue
            if (isinstance(va, float) and isinstance(vb, float)
                    and np.isnan(va) and np.isnan(vb)):
                continue
            nome = a[0][j] if a and j < len(a[0]) else f'col{j}'
            problemas.append(f"{caso}: linha {i}, coluna {nome!r}: "
                             f"to_excel={va!r} != escrever_aba={vb!r}")
            if len(problemas) > 20:
                return problemas
    return problemas


class TestEscreverAbaEquivaleAoToExcel(unittest.TestCase):

    def _escrever_pelos_dois_caminhos(self, df, nome_aba, formatar=False, sufixo=''):
        """Grava df pelo caminho antigo e pelo novo; devolve as duas matrizes de células."""
        import tempfile
        dir_tmp = tempfile.mkdtemp()
        p_antigo = os.path.join(dir_tmp, f'antigo{sufixo}.xlsx')
        p_novo = os.path.join(dir_tmp, f'novo{sufixo}.xlsx')
        opcoes = {'options': {'strings_to_formulas': False}}

        w = pd.ExcelWriter(p_antigo, engine='xlsxwriter', engine_kwargs=opcoes)
        df.to_excel(w, sheet_name=nome_aba, index=False)
        if formatar:
            ggci.aplicar_formatacao_visual(w, nome_aba, df)
        w.close()

        w = pd.ExcelWriter(p_novo, engine='xlsxwriter', engine_kwargs=opcoes)
        ggci.escrever_aba(w, nome_aba, df)
        if formatar:
            ggci.aplicar_formatacao_visual(w, nome_aba, df)
        w.close()

        return celulas(p_antigo, nome_aba), celulas(p_novo, nome_aba)

    def test_todos_os_dtypes_do_relatorio(self):
        df = pd.DataFrame({
            'int64_puro':     pd.Series([1, 2, 3, 4, 5, 6], dtype='int64'),
            'int64_nullable': pd.Series([10, pd.NA, 30, pd.NA, 50, 0], dtype='Int64'),
            'float_com_nan':  [1.5, np.nan, -3.25, 0.0, 1e12, -0.0001],
            'float_inf':      [np.inf, -np.inf, np.nan, 1.0, 2.0, 3.0],
            'texto':          ['Ana', 'José da Silva', '', 'ÁÉÍÓÚ ção', 'x' * 300, 'Nao Processado'],
            'object_misto':   ['abc', 123, None, 45.6, pd.NA, True],
            'traco':          ['-', '-', 'CPF 123', '-', 'IES Federal', '-'],
            'datetime':       pd.to_datetime(['2026-01-15', '2026-08-11', None,
                                              '2025-12-31', '2026-02-28', '2024-06-01']),
            'bool_puro':      [True, False, True, False, True, False],
            'parece_formula': ['=SOMA(A1:A2)', '+55 62 9999', '-10', '@usuario', 'ok', '=1+1'],
            'zero_e_vazio':   [0, 0.0, '', None, np.nan, 0],
        })
        a, b = self._escrever_pelos_dois_caminhos(df, 'Contrato')
        problemas = comparar_celulas('dtypes', a, b)
        self.assertEqual([], problemas, '\n'.join(problemas))

    def test_dataframe_vazio_com_aviso(self):
        df = pd.DataFrame([{"Aviso": "Nenhum documento encontrado ou processado para este tipo"}])
        a, b = self._escrever_pelos_dois_caminhos(df, 'Financiamento', formatar=True)
        problemas = comparar_celulas('aviso', a, b)
        self.assertEqual([], problemas, '\n'.join(problemas))

    @unittest.skipUnless(os.path.exists(BASELINE),
                         'baseline ausente — gere o relatório e copie para tests/baseline/')
    def test_dados_reais_do_baseline_com_formatacao(self):
        """Passa cada aba real do baseline pelos dois caminhos, com a formatação de verdade."""
        wb = openpyxl.load_workbook(BASELINE, read_only=True)
        abas = wb.sheetnames
        wb.close()

        falhas = []
        for aba in abas:
            linhas = []
            wb = openpyxl.load_workbook(BASELINE, read_only=True)
            for n, linha in enumerate(wb[aba].iter_rows(values_only=True)):
                linhas.append(linha)
                if n >= LINHAS_AMOSTRA:
                    break
            wb.close()
            if len(linhas) < 2:
                continue

            cabecalho = [str(c) if c is not None else f'col{i}' for i, c in enumerate(linhas[0])]
            df = pd.DataFrame(linhas[1:], columns=cabecalho)
            a, b = self._escrever_pelos_dois_caminhos(df, aba, formatar=True,
                                                      sufixo=str(abs(hash(aba))))
            falhas += comparar_celulas(f'[{aba}]', a, b)

        self.assertEqual([], falhas, f'{len(falhas)} divergências:\n' + '\n'.join(falhas[:20]))


if __name__ == '__main__':
    unittest.main()
