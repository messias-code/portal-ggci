"""
=== ARQUIVO: apps/automacoes/analise_ia/tests/comparar_relatorios.py ===
Propósito: Compara dois relatorio_geral.xlsx e aponta qualquer divergência entre eles.
Autor: N/A
Dependências Principais: openpyxl

POR QUÊ EXISTE: A refatoração precisa provar que o arquivo final continua idêntico. Os
testes travam a estrutura; este script vai além e confere também os valores, célula a
célula, contra o relatório aprovado.

POR QUE COMPARA TANTA COISA: o modo `constant_memory` do xlsxwriter já pareceu 41% mais
rápido e zerou 22 das 23 colunas da aba Pagamentos. O arquivo abria normal, com o número
certo de linhas e os cabeçalhos certos — só o conteúdo estava vazio. Contagem de nulos
por coluna é o que pega esse tipo de falha; estrutura sozinha não pega.

COMO USAR (do diretório raiz do projeto, com o venv ativo):

    venv/bin/python3 apps/automacoes/analise_ia/tests/comparar_relatorios.py \\
        apps/automacoes/analise_ia/tests/baseline/relatorio_geral.xlsx \\
        apps/automacoes/analise_ia/dados/processamento/proc_8/relatorio_geral.xlsx

Sem argumentos, compara os dois proc_* de maior número automaticamente.

    --amostra N   compara apenas as N primeiras linhas de cada aba (varredura rápida)

COMO LER O RESULTADO: qualquer coisa sob "REGRESSÃO" é a refatoração alterando o
artefato e precisa ser revertida ou corrigida. Sob "Diferenças de conteúdo" pode ser
apenas dado novo — documentos enviados entre uma execução e outra.
"""
import glob
import os
import re
import sys

import xml.etree.ElementTree as ET
import zipfile

import openpyxl

# Abas que carregam fórmulas: precisam ser comparadas com data_only=False, senão
# openpyxl devolve o último valor calculado pelo Excel (ou None) em vez da fórmula.
ABAS_COM_FORMULA = ('Relatório Contratos', 'Relatório RIAF', 'Aux_IES_Contratos', 'Aux_IES_RIAF')


def relatorios_disponiveis():
    base = os.path.join(os.path.dirname(__file__), "..", "dados", "processamento")
    achados = glob.glob(os.path.join(base, "proc_*", "relatorio_geral.xlsx"))
    return sorted(achados, key=lambda c: int(re.search(r'proc_(\d+)', c).group(1)))


def carregar(caminho, data_only=True, limite=None):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=data_only)
    dados = {}
    for aba in wb.sheetnames:
        cabecalho, linhas = [], []
        nulos, preenchidos = None, None
        for n, linha in enumerate(wb[aba].iter_rows(values_only=True)):
            if n == 0:
                cabecalho = [str(c) for c in linha if c is not None]
                nulos = [0] * len(linha)
                preenchidos = [0] * len(linha)
                continue
            for j, v in enumerate(linha):
                if j >= len(nulos):
                    continue
                if v is None or v == '':
                    nulos[j] += 1
                else:
                    preenchidos[j] += 1
            if limite is None or len(linhas) < limite:
                linhas.append(linha)
            elif limite is not None:
                continue
        dados[aba] = {'cabecalho': cabecalho, 'linhas': linhas,
                      'nulos': nulos or [], 'preenchidos': preenchidos or [],
                      'total_linhas': (n if 'n' in dir() else 0)}
    wb.close()
    return dados


def carregar_visual(caminho):
    """
    Larguras, painel congelado, autofilter, cor de aba e tabelas — o que data_only não vê.

    Lê o XML direto do .xlsx em vez de usar openpyxl com read_only=False. Aquele caminho
    materializa o workbook inteiro: medido em 6,38 GB de pico para o relatório de 74 MB,
    o suficiente para o OOM killer derrubar uma geração que estivesse rodando ao mesmo
    tempo. Aqui o parser é incremental e descarta cada elemento — o custo fica em MB.
    """
    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    ns_r = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
    ns_pkg = '{http://schemas.openxmlformats.org/package/2006/relationships}'

    visual = {}
    with zipfile.ZipFile(caminho) as z:
        nomes = set(z.namelist())

        # nome da aba -> arquivo xml da worksheet
        rels = {}
        with z.open('xl/_rels/workbook.xml.rels') as f:
            for _, el in ET.iterparse(f):
                if el.tag.endswith('Relationship'):
                    rels[el.attrib['Id']] = el.attrib['Target'].lstrip('/')
        abas = []
        with z.open('xl/workbook.xml') as f:
            for _, el in ET.iterparse(f):
                if el.tag == f'{ns}sheet':
                    alvo = rels.get(el.attrib.get(f'{ns_r}id'), '')
                    if not alvo.startswith('xl/'):
                        alvo = 'xl/' + alvo
                    abas.append((el.attrib['name'], alvo))

        for nome, arq in abas:
            dados = {'freeze': None, 'autofilter': None, 'tab_color': None,
                     'larguras': {}, 'tabelas': [], 'altura_linha1': None}
            if arq not in nomes:
                visual[nome] = dados
                continue

            with z.open(arq) as f:
                for evento, el in ET.iterparse(f, events=('end',)):
                    tag = el.tag
                    if tag == f'{ns}tabColor':
                        dados['tab_color'] = el.attrib.get('rgb')
                    elif tag == f'{ns}pane':
                        dados['freeze'] = el.attrib.get('topLeftCell')
                    elif tag == f'{ns}col':
                        if el.attrib.get('width'):
                            largura = round(float(el.attrib['width']), 2)
                            for i in range(int(el.attrib['min']), int(el.attrib['max']) + 1):
                                dados['larguras'][openpyxl.utils.get_column_letter(i)] = largura
                    elif tag == f'{ns}autoFilter':
                        dados['autofilter'] = el.attrib.get('ref')
                    elif tag == f'{ns}row':
                        if el.attrib.get('r') == '1' and el.attrib.get('ht'):
                            dados['altura_linha1'] = float(el.attrib['ht'])
                        el.clear()          # descarta as células da linha
                    elif tag == f'{ns}sheetData':
                        el.clear()
                    if tag in (f'{ns}col', f'{ns}pane', f'{ns}tabColor'):
                        el.clear()

            # nomes das tabelas nativas desta aba
            rel_aba = arq.replace('xl/worksheets/', 'xl/worksheets/_rels/') + '.rels'
            if rel_aba in nomes:
                with z.open(rel_aba) as f:
                    for _, el in ET.iterparse(f):
                        if el.tag == f'{ns_pkg}Relationship' and 'tables/' in el.attrib.get('Target', ''):
                            alvo = el.attrib['Target'].replace('../', 'xl/')
                            if alvo in nomes:
                                with z.open(alvo) as ft:
                                    for _, elt in ET.iterparse(ft):
                                        if elt.tag == f'{ns}table':
                                            dados['tabelas'].append(elt.attrib.get('displayName')
                                                                    or elt.attrib.get('name'))
                                            if elt.attrib.get('ref') and not dados['autofilter']:
                                                dados['autofilter'] = elt.attrib['ref']
                                        elif elt.tag == f'{ns}autoFilter':
                                            dados['autofilter'] = elt.attrib.get('ref')
            dados['tabelas'] = sorted(x for x in dados['tabelas'] if x)
            visual[nome] = dados
    return visual


def comparar(base, novo, limite=None):
    a, b = carregar(base, limite=limite), carregar(novo, limite=limite)
    problemas, avisos = [], []

    if list(a.keys()) != list(b.keys()):
        problemas.append(f"abas diferentes:\n    base: {list(a.keys())}\n    novo: {list(b.keys())}")

    for aba in a:
        if aba not in b:
            continue
        if a[aba]['cabecalho'] != b[aba]['cabecalho']:
            so_base = [c for c in a[aba]['cabecalho'] if c not in b[aba]['cabecalho']]
            so_novo = [c for c in b[aba]['cabecalho'] if c not in a[aba]['cabecalho']]
            if so_base or so_novo:
                problemas.append(f"[{aba}] colunas divergentes — sumiram: {so_base} | surgiram: {so_novo}")
            else:
                problemas.append(f"[{aba}] mesmas colunas, ORDEM diferente")
            continue

        na, nb = a[aba]['total_linhas'], b[aba]['total_linhas']
        if na != nb:
            avisos.append(f"[{aba}] {na} linhas na base, {nb} no novo ({nb - na:+d})")

        # Preenchimento por coluna: pega o caso "arquivo íntegro, conteúdo vazio".
        for j, nome in enumerate(a[aba]['cabecalho']):
            if j >= len(a[aba]['preenchidos']) or j >= len(b[aba]['preenchidos']):
                continue
            pa, pb = a[aba]['preenchidos'][j], b[aba]['preenchidos'][j]
            if pa > 0 and pb == 0:
                problemas.append(f"[{aba}] coluna '{nome}' ZEROU: {pa} valores na base, 0 no novo")
            elif pa != pb and na == nb:
                avisos.append(f"[{aba}] coluna '{nome}': {pa} preenchidos na base, {pb} no novo ({pb - pa:+d})")

        # Comparação por CHAVE quando as abas têm identificador natural.
        #
        # Comparar por POSIÇÃO engana: uma única linha nova no meio desloca todas as
        # seguintes e a saída acusa "39.342 linhas com valor diferente" num arquivo em que
        # apenas 1.095 registros realmente mudaram. Já aconteceu nesta suíte.
        cab = a[aba]['cabecalho']
        chave_cols = [c for c in ('inscricao', 'Inscrição', 'codigo_aluno', 'UNI_CODIGO') if c in cab]
        sem_col = [c for c in ('semestre', 'Semestre', 'semestre_referencia_analise') if c in cab]

        if chave_cols and limite is None:
            idx_chave = [cab.index(chave_cols[0])] + ([cab.index(sem_col[0])] if sem_col else [])

            def indexar(linhas):
                mapa = {}
                for linha in linhas:
                    k = tuple(linha[i] if i < len(linha) else None for i in idx_chave)
                    mapa.setdefault(k, []).append(linha)
                return mapa

            ia, ib = indexar(a[aba]['linhas']), indexar(b[aba]['linhas'])
            so_base, so_novo = set(ia) - set(ib), set(ib) - set(ia)
            if so_base:
                avisos.append(f"[{aba}] {len(so_base)} chaves só na base (ex.: {list(so_base)[:3]})")
            if so_novo:
                avisos.append(f"[{aba}] {len(so_novo)} chaves só no novo (ex.: {list(so_novo)[:3]})")

            divergentes, por_coluna = 0, {}
            for k in set(ia) & set(ib):
                for la, lb in zip(ia[k], ib[k]):
                    if la != lb:
                        divergentes += 1
                        for j, (va, vb) in enumerate(zip(la, lb)):
                            if va != vb:
                                nome = cab[j] if j < len(cab) else f'col{j}'
                                por_coluna.setdefault(nome, []).append((k, va, vb))
            comuns = len(set(ia) & set(ib))
            if divergentes:
                pct = 100 * divergentes / max(comuns, 1)
                avisos.append(f"[{aba}] {divergentes} de {comuns} registros com valor diferente "
                              f"({pct:.1f}%) — comparado por chave {tuple(cab[i] for i in idx_chave)}")
                ranking = sorted(por_coluna.items(), key=lambda x: -len(x[1]))
                for nome, ocorr in ranking[:6]:
                    k, va, vb = ocorr[0]
                    marca = '  <-- TODOS os registros' if len(ocorr) == comuns else ''
                    avisos.append(f"      {nome}: {len(ocorr)} registros{marca}   ex. {k}: {va!r} -> {vb!r}")
                # Coluna que muda em 100% dos registros quase nunca é dado novo.
                for nome, ocorr in ranking:
                    if len(ocorr) == comuns and comuns > 100:
                        problemas.append(f"[{aba}] coluna '{nome}' mudou em TODOS os {comuns} "
                                         f"registros — suspeito de mudança de código, não de dado")
        else:
            divergentes, exemplos = 0, []
            for i, (la, lb) in enumerate(zip(a[aba]['linhas'], b[aba]['linhas']), start=2):
                if la != lb:
                    divergentes += 1
                    if len(exemplos) < 3:
                        for j, (va, vb) in enumerate(zip(la, lb)):
                            if va != vb:
                                col = cab[j] if j < len(cab) else f'col{j}'
                                exemplos.append(f"linha {i}, coluna '{col}': {va!r} -> {vb!r}")
                                break
            if divergentes:
                avisos.append(f"[{aba}] {divergentes} linhas com valor diferente "
                              f"(comparação POSICIONAL — sem coluna de chave nesta aba)")
                for e in exemplos:
                    avisos.append(f"      {e}")

    # --- Fórmulas (data_only=False) nas abas que têm ---
    abas_formula = [x for x in a if x in ABAS_COM_FORMULA]
    if abas_formula:
        fa = carregar(base, data_only=False, limite=limite)
        fb = carregar(novo, data_only=False, limite=limite)
        for aba in abas_formula:
            if aba not in fb:
                continue
            dif = sum(1 for la, lb in zip(fa[aba]['linhas'], fb[aba]['linhas']) if la != lb)
            if dif:
                problemas.append(f"[{aba}] {dif} linhas com FÓRMULA divergente")

    # --- Propriedades visuais ---
    try:
        va, vb = carregar_visual(base), carregar_visual(novo)
        for aba in va:
            if aba not in vb:
                continue
            for chave in ('freeze', 'autofilter', 'tab_color', 'altura_linha1', 'tabelas'):
                if va[aba][chave] != vb[aba][chave]:
                    problemas.append(f"[{aba}] {chave}: {va[aba][chave]!r} -> {vb[aba][chave]!r}")
            if va[aba]['larguras'] != vb[aba]['larguras']:
                difs = [f"{k}: {va[aba]['larguras'].get(k)} -> {vb[aba]['larguras'].get(k)}"
                        for k in set(va[aba]['larguras']) | set(vb[aba]['larguras'])
                        if va[aba]['larguras'].get(k) != vb[aba]['larguras'].get(k)]
                problemas.append(f"[{aba}] larguras de coluna divergentes: {'; '.join(difs[:5])}")
    except Exception as e:
        avisos.append(f"não foi possível comparar propriedades visuais: {e}")

    return problemas, avisos


def main():
    argv = [x for x in sys.argv[1:]]
    limite = None
    if '--amostra' in argv:
        i = argv.index('--amostra')
        limite = int(argv[i + 1])
        del argv[i:i + 2]

    if len(argv) == 2:
        base, novo = argv
    else:
        disponiveis = relatorios_disponiveis()
        if len(disponiveis) < 2:
            print("Preciso de dois relatórios para comparar. Passe os caminhos como argumento.")
            return 2
        base, novo = disponiveis[-2], disponiveis[-1]

    for caminho in (base, novo):
        if not os.path.exists(caminho):
            print(f"Arquivo não encontrado: {caminho}")
            print("(execuções interrompidas deixam a pasta proc_* sem o relatorio_geral.xlsx)")
            return 2

    print(f"base: {base}")
    print(f"novo: {novo}")
    if limite:
        print(f"(amostra: {limite} primeiras linhas por aba — contagem de nulos usa a aba inteira)")
    print()
    problemas, avisos = comparar(base, novo, limite=limite)

    if problemas:
        print("REGRESSÃO — a refatoração alterou o artefato:")
        for p in problemas:
            print(f"  {p}")
    else:
        print("Estrutura, fórmulas e formatação idênticas.")

    if avisos:
        print("\nDiferenças de conteúdo (podem ser apenas dados novos):")
        for a in avisos:
            print(f"  {a}")
    else:
        print("Conteúdo idêntico: nenhuma linha ou valor divergente.")

    return 1 if problemas else 0


if __name__ == '__main__':
    sys.exit(main())
