"""
=== ARQUIVO: apps/automacoes/analise_ia/tests/comparar_csv_xlsx.py ===
Propósito: Confere se o ZIP de CSV traz o mesmo conteúdo das abas do relatorio_geral.xlsx.
Autor: N/A
Dependências Principais: pandas, openpyxl, zipfile

POR QUÊ EXISTE: os dois formatos saem da mesma função (`montar_abas_de_dados`), mas o CSV
já divergiu do XLSX no passado — mantinha `tipo_documento`, ordenava sem o semestre e não
normalizava Pagamentos. Este script prova a equivalência aba por aba.

As 4 abas de relatório gerencial (Relatório Contratos, Aux_IES_Contratos, Relatório RIAF,
Aux_IES_RIAF) são ignoradas de propósito: dependem de fórmulas do Excel e não têm
equivalente em texto puro.

COMO USAR (do diretório raiz do projeto):

    venv/bin/python3 apps/automacoes/analise_ia/tests/comparar_csv_xlsx.py \\
        <caminho>/relatorio_geral.xlsx <caminho>/relatorio_geral.zip

COMO LER: "FALTA NO ZIP" pode ser configuração (as flags gerar_pagamentos /
gerar_quantitativo desmarcadas removem as abas nos DOIS formatos) — confira o payload
antes de tratar como defeito. Divergência de coluna ou de valor é defeito.
"""
import io
import os
import sys
import zipfile

import openpyxl
import pandas as pd

ABAS_SO_DO_EXCEL = ('Relatório Contratos', 'Aux_IES_Contratos', 'Relatório RIAF', 'Aux_IES_RIAF')
CHAVES = ('inscricao', 'Inscrição', 'codigo_aluno', 'UNI_CODIGO')
SEMESTRES = ('semestre', 'Semestre', 'semestre_referencia_analise')


def ler_abas_xlsx(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    abas = {}
    for nome in wb.sheetnames:
        if nome in ABAS_SO_DO_EXCEL:
            continue
        it = wb[nome].iter_rows(values_only=True)
        try:
            cabecalho = [str(c) if c is not None else '' for c in next(it)]
        except StopIteration:
            continue
        abas[nome] = (cabecalho, list(it))
    wb.close()
    return abas


def ler_csvs_do_zip(caminho):
    abas = {}
    with zipfile.ZipFile(caminho) as z:
        for info in z.infolist():
            if not info.filename.lower().endswith('.csv'):
                continue
            nome = os.path.splitext(info.filename)[0]
            bruto = z.read(info).decode('utf-8-sig')
            df = pd.read_csv(io.StringIO(bruto), sep=';', dtype=str, keep_default_na=False,
                             low_memory=False)
            abas[nome] = ([str(c) for c in df.columns], df)
    return abas


def normalizar(v):
    """Aproxima a leitura do CSV (tudo texto) da do Excel (tipado)."""
    if v is None:
        return ''
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else f'{v:.10g}'
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def equivalentes(vx, vc):
    """
    Compara uma célula do xlsx com a mesma célula do csv.

    Número precisa ser comparado como número, não como texto: o Excel guarda o double e
    devolve 3353.22, enquanto o `to_csv` do pandas escreve o repr completo do mesmo double,
    3353.2200000000003. São o mesmo valor — comparar como string acusa divergência em
    dezenas de milhares de células que estão corretas.
    """
    a, b = normalizar(vx), normalizar(vc)
    if a == b:
        return True
    # Data à meia-noite: o Excel guarda o datetime e devolve '2025-01-13 00:00:00',
    # o to_csv escreve '2025-01-13'. É o mesmo instante, só a representação muda.
    if a.replace(' 00:00:00', '') == b.replace(' 00:00:00', ''):
        return True
    try:
        fa, fb = float(a.replace(',', '.')), float(b.replace(',', '.'))
    except (ValueError, AttributeError):
        return False
    if fa == fb:
        return True
    escala = max(abs(fa), abs(fb), 1.0)
    return abs(fa - fb) <= 1e-9 * escala


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    caminho_xlsx, caminho_zip = sys.argv[1], sys.argv[2]
    for c in (caminho_xlsx, caminho_zip):
        if not os.path.exists(c):
            print(f'Arquivo não encontrado: {c}')
            return 2

    print(f'xlsx: {caminho_xlsx}')
    print(f'zip : {caminho_zip}\n')

    abas_x = ler_abas_xlsx(caminho_xlsx)
    abas_c = ler_csvs_do_zip(caminho_zip)

    print(f'abas comparáveis no xlsx ({len(abas_x)}): {list(abas_x)}')
    print(f'csvs no zip ({len(abas_c)}): {list(abas_c)}\n')

    falta_no_zip = [a for a in abas_x if a not in abas_c]
    sobra_no_zip = [a for a in abas_c if a not in abas_x]
    if falta_no_zip:
        print(f'FALTA NO ZIP: {falta_no_zip}')
    if sobra_no_zip:
        print(f'SOBRA NO ZIP: {sobra_no_zip}')
    if not falta_no_zip and not sobra_no_zip:
        print('Conjunto de abas idêntico.')
    print()

    problemas = 0
    for nome in abas_x:
        if nome not in abas_c:
            continue
        cab_x, linhas_x = abas_x[nome]
        cab_c, df_c = abas_c[nome]

        print(f'--- {nome}')
        if cab_x != cab_c:
            so_x = [c for c in cab_x if c not in cab_c]
            so_c = [c for c in cab_c if c not in cab_x]
            if so_x or so_c:
                print(f'    COLUNAS DIVERGEM — só no xlsx: {so_x} | só no csv: {so_c}')
            else:
                print(f'    mesmas colunas, ORDEM diferente')
                print(f'      xlsx: {cab_x[:6]}')
                print(f'      csv : {cab_c[:6]}')
            problemas += 1
            continue
        print(f'    colunas: {len(cab_x)} idênticas e na mesma ordem')

        if len(linhas_x) != len(df_c):
            print(f'    LINHAS: {len(linhas_x)} no xlsx, {len(df_c)} no csv ({len(df_c)-len(linhas_x):+d})')
            problemas += 1
        else:
            print(f'    linhas: {len(linhas_x)} em ambos')

        # Ordem das linhas: compara a coluna de chave posição por posição.
        chave = next((c for c in CHAVES if c in cab_x), None)
        if chave and len(linhas_x) == len(df_c):
            i = cab_x.index(chave)
            seq_x = [normalizar(l[i]) for l in linhas_x]
            seq_c = [normalizar(v) for v in df_c[chave].tolist()]
            if seq_x == seq_c:
                print(f'    ordem das linhas: idêntica (por {chave})')
            else:
                fora = sum(1 for a, b in zip(seq_x, seq_c) if a != b)
                print(f'    ORDEM DAS LINHAS DIVERGE em {fora} posições (por {chave})')
                problemas += 1

        # Valores: amostra ampla comparando célula a célula na mesma posição.
        if len(linhas_x) == len(df_c):
            limite = min(len(linhas_x), 5000)
            difs, exemplos = 0, []
            valores_c = df_c.head(limite).values
            for r in range(limite):
                for j in range(len(cab_x)):
                    if not equivalentes(linhas_x[r][j], valores_c[r][j]):
                        difs += 1
                        if len(exemplos) < 4:
                            exemplos.append(f'linha {r+2}, {cab_x[j]!r}: '
                                            f'xlsx={normalizar(linhas_x[r][j])!r} '
                                            f'csv={normalizar(valores_c[r][j])!r}')
            if difs:
                print(f'    VALORES: {difs} células divergentes nas {limite} primeiras linhas')
                for e in exemplos:
                    print(f'        {e}')
                problemas += 1
            else:
                print(f'    valores: idênticos nas {limite} primeiras linhas x {len(cab_x)} colunas')
        print()

    print('=' * 70)
    if problemas == 0 and not falta_no_zip and not sobra_no_zip:
        print('CSV equivalente ao XLSX em todas as abas comparáveis.')
        return 0
    print(f'{problemas} aba(s) com divergência de estrutura/valor.')
    if falta_no_zip:
        print(f'Abas ausentes no ZIP: {falta_no_zip}')
        print('(confira as flags gerar_pagamentos / gerar_quantitativo no payload antes de '
              'tratar como defeito)')
    return 1


if __name__ == '__main__':
    sys.exit(main())
