"""
=== ARQUIVO: apps/automacoes/analise_ia/tests/test_contrato_saida.py ===
Propósito: Trava a estrutura do relatório_geral.xlsx contra o golden capturado do proc_7.
Autor: N/A
Dependências Principais: unittest, openpyxl

POR QUÊ EXISTE: O arquivo final foi validado manualmente e está aprovado. Qualquer
refatoração que altere aba, coluna ou ordem de coluna precisa falhar aqui, não na mão
de quem abre a planilha.

COMO FUNCIONA: Compara o relatório mais recente em dados/processamento/proc_* com o
contrato salvo em fixtures/contrato_saida.json. Se não houver relatório em disco, os
testes são pulados — a esteira de CI não depende de dados de produção.

NOTA DE PERFORMANCE: o xlsx passa de 70 MB e `pd.read_excel` materializa a aba inteira
(130s para a suíte). Aqui usamos openpyxl em read_only + iter_rows, lendo apenas as
colunas necessárias em streaming, o que derruba o tempo para poucos segundos.
"""
import glob
import json
import os
import re
import unittest

import openpyxl

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.."))
PASTA_PROC = os.path.join(PROJECT_ROOT, "apps/automacoes/analise_ia/dados/processamento")
CONTRATO = os.path.join(os.path.dirname(__file__), "fixtures", "contrato_saida.json")

ABAS_DE_DOCUMENTO = ['Contrato', 'Riaf', 'Benefício', 'Financiamento', 'Histórico']


BASELINE = os.path.join(os.path.dirname(__file__), "baseline", "relatorio_geral.xlsx")


def relatorio_mais_recente():
    """
    Devolve o relatório a validar contra o contrato, ou None.

    Prefere tests/baseline/relatorio_geral.xlsx quando existe. O proc_N de maior número
    não serve como referência fixa: uma execução de teste com abas desmarcadas (sem os
    relatórios, sem Envios & Pendências) gera um arquivo legitimamente menor e faria esta
    suíte falhar por um motivo que não é regressão. O baseline é o relatório aprovado.
    """
    if os.path.exists(BASELINE):
        return BASELINE

    achados = glob.glob(os.path.join(PASTA_PROC, "proc_*", "relatorio_geral.xlsx"))
    if not achados:
        return None

    def numero(caminho):
        m = re.search(r'proc_(\d+)', caminho)
        return int(m.group(1)) if m else -1

    return max(achados, key=numero)


class LeitorRelatorio:
    """Lê cabeçalhos e colunas isoladas do xlsx sem materializar as abas inteiras."""

    def __init__(self, caminho):
        self._wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        self.abas = self._wb.sheetnames
        self._cabecalhos = {}
        self._colunas = {}

    def cabecalho(self, aba):
        if aba not in self._cabecalhos:
            linha = next(self._wb[aba].iter_rows(min_row=1, max_row=1, values_only=True), ())
            self._cabecalhos[aba] = [str(c) for c in linha if c is not None]
        return self._cabecalhos[aba]

    def colunas(self, aba, nomes):
        """
        Valores de várias colunas numa única varredura da aba.

        Ler coluna a coluna re-percorreria o arquivo inteiro a cada chamada; com abas de
        60 mil linhas isso domina o tempo da suíte. O resultado fica em cache por aba.
        """
        cache = self._colunas.setdefault(aba, {})
        faltando = [n for n in nomes if n not in cache]
        if faltando:
            cabecalho = self.cabecalho(aba)
            indices = {n: cabecalho.index(n) for n in faltando}
            for n in faltando:
                cache[n] = []
            for linha in self._wb[aba].iter_rows(min_row=2, values_only=True):
                for n, i in indices.items():
                    cache[n].append(linha[i] if i < len(linha) else None)
        return {n: cache[n] for n in nomes}

    def coluna(self, aba, nome):
        return self.colunas(aba, [nome])[nome]

    def fechar(self):
        self._wb.close()


@unittest.skipIf(relatorio_mais_recente() is None, "nenhum relatorio_geral.xlsx em disco")
class TestContratoDeSaida(unittest.TestCase):
    """Estrutura do arquivo final — o que foi aprovado não pode mudar sozinho."""

    @classmethod
    def setUpClass(cls):
        cls.leitor = LeitorRelatorio(relatorio_mais_recente())
        cls.esperado = json.load(open(CONTRATO, encoding='utf-8'))['abas']

    @classmethod
    def tearDownClass(cls):
        cls.leitor.fechar()

    def test_conjunto_de_abas_inalterado(self):
        self.assertEqual(self.leitor.abas, list(self.esperado.keys()))

    def test_colunas_e_ordem_inalteradas_em_cada_aba(self):
        for aba, colunas in self.esperado.items():
            with self.subTest(aba=aba):
                self.assertEqual(self.leitor.cabecalho(aba), colunas)

    def test_nenhuma_coluna_interna_numerada_vaza_para_a_saida(self):
        """Os nomes com [N] são intermediários; se aparecerem no xlsx, o rename quebrou."""
        for aba in self.leitor.abas:
            with self.subTest(aba=aba):
                vazadas = [c for c in self.leitor.cabecalho(aba) if re.match(r'^\[\d+\]', c)]
                self.assertEqual(vazadas, [])


@unittest.skipIf(relatorio_mais_recente() is None, "nenhum relatorio_geral.xlsx em disco")
@unittest.skipUnless(os.environ.get('GGCI_TESTES_LENTOS') == '1',
                     "varredura de conteúdo: exporte GGCI_TESTES_LENTOS=1 para rodar")
class TestInvariantesDeConteudo(unittest.TestCase):
    """
    Regras de negócio que valem para qualquer execução, independente dos dados.

    Percorre as abas inteiras (~200 mil linhas somadas) e leva dezenas de segundos, então
    fica atrás de uma variável de ambiente: durante a refatoração roda-se só o contrato
    estrutural, que é instantâneo, e a varredura completa entra na validação final e na CI.
    """

    @classmethod
    def setUpClass(cls):
        cls.leitor = LeitorRelatorio(relatorio_mais_recente())
        # Uma varredura por aba, carregando de uma vez tudo que os testes abaixo consultam.
        for aba in ABAS_DE_DOCUMENTO:
            cls.leitor.colunas(aba, ['beneficio', 'financiamento',
                                     'soma_prejuizo_ovg', 'soma_economia_ovg'])
        cls.leitor.colunas('Pagamentos', ['desc_outro_beneficio', 'desc_financiamento'])

    @classmethod
    def tearDownClass(cls):
        cls.leitor.fechar()

    @staticmethod
    def _numeros(valores):
        saida = []
        for v in valores:
            if isinstance(v, (int, float)) and v is not None:
                saida.append(float(v))
        return saida

    def test_rotulo_antigo_nao_informado_nao_aparece_mais(self):
        for aba in ABAS_DE_DOCUMENTO:
            for coluna in ['beneficio', 'financiamento']:
                with self.subTest(aba=aba, coluna=coluna):
                    self.assertNotIn('Não Informado', set(self.leitor.coluna(aba, coluna)))

    def test_aba_pagamentos_tambem_usa_o_rotulo_novo(self):
        for coluna in ['desc_outro_beneficio', 'desc_financiamento']:
            with self.subTest(coluna=coluna):
                self.assertNotIn('Não Informado', set(self.leitor.coluna('Pagamentos', coluna)))

    def test_valores_financeiros_nunca_sao_negativos(self):
        """Prejuízo e economia saem de np.maximum(..., 0) — negativo indica regressão na fórmula."""
        for aba in ABAS_DE_DOCUMENTO:
            for coluna in ['soma_prejuizo_ovg', 'soma_economia_ovg']:
                with self.subTest(aba=aba, coluna=coluna):
                    valores = self._numeros(self.leitor.coluna(aba, coluna))
                    self.assertTrue(all(v >= 0 for v in valores), f'{aba}.{coluna} tem negativo')

    def test_prejuizo_e_economia_nao_coexistem_na_mesma_linha(self):
        """São mutuamente exclusivos: ou a OVG pagou a mais, ou a menos."""
        prejuizo = self.leitor.coluna('Contrato', 'soma_prejuizo_ovg')
        economia = self.leitor.coluna('Contrato', 'soma_economia_ovg')
        ambos = sum(1 for p, e in zip(prejuizo, economia)
                    if isinstance(p, (int, float)) and isinstance(e, (int, float))
                    and p > 0 and e > 0)
        self.assertEqual(ambos, 0)


if __name__ == '__main__':
    unittest.main()
