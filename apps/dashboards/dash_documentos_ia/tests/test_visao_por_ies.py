"""
=== ARQUIVO: apps/dashboards/dash_documentos_ia/tests/test_visao_por_ies.py ===
Propósito: Trava a visão por IES — o resumo por instituição e as peças que a tela usa.
Autor: N/A
Dependências Principais: pandas, django.test

POR QUÊ EXISTE: a vista por IES conta os MESMOS SEIS BALDES da vista de beneficiários,
por outro eixo. É uma segunda contagem sobre a mesma regra, e é aí que mora o risco: se
`_resumo_por_ies` classificasse de um jeito e `_balde_do_documento` de outro, a coluna
"Pendentes" daqui não bateria com a fatia "Pendentes" de lá — as duas na mesma tela, sem
erro em lugar nenhum, e sem como saber qual está certa.

O que este arquivo garante:

  1. Os seis baldes COBREM todas as linhas (a soma dos totais é o tamanho do recorte).
  2. Beneficiário é PESSOA: o mesmo CPF em cinco documentos conta uma vez, e o total
     geral não é a soma das colunas — quem muda de IES no meio do período aparece nas
     duas.
  3. O filtro de documento recorta de verdade, e rótulo desconhecido não devolve a base
     inteira por engano.
  4. As peças que o JavaScript procura na página existem, e o filtro de documento nasce
     recolhido — ele só vale no modo IES.
"""
import pandas as pd
from django.test import Client, SimpleTestCase, TestCase
from django.urls import reverse

from apps.inicio.gestao_acessos.models import Usuario

from apps.dashboards.dash_documentos_ia import views


def _abas(**colunas):
    """
    DataFrame no formato que `_carregar_abas` devolve — já com `documento` e com
    `status_ia` em caixa alta, que é como o resto do arquivo compara.
    """
    df = pd.DataFrame(colunas)
    df['status_ia'] = df['status_ia'].astype('string').str.upper()
    df['faculdade'] = df['faculdade'].astype('string')
    return df


class Pedido:
    """`request` mínimo: as funções deste módulo só leem `request.GET`."""

    def __init__(self, **parametros):
        self.GET = parametros


class TestResumoPorIES(SimpleTestCase):
    """A contagem por instituição, sem Django e sem disco."""

    def setUp(self):
        # Duas IES, três pessoas. A pessoa 3 aparece nas DUAS instituições — mudou de
        # IES no meio do período, que é o caso que impede somar as colunas.
        self.df = _abas(
            faculdade=['IES A', 'IES A', 'IES A', 'IES B', 'IES B'],
            cpf=['1', '1', '2', '3', '3'],
            documento=['CONTRATO', 'RIAF', 'CONTRATO', 'CONTRATO', 'RIAF'],
            status_ia=['Válido', 'Ausente', 'Não Processado', 'Inadimplente', 'Válido'],
            documento_ausente=['NÃO', 'NÃO', 'NÃO', 'SIM', 'NÃO'],
            veredito_documento=['VÁLIDO', '', '', '', 'VÁLIDO'],
        )

    def test_os_seis_baldes_cobrem_todas_as_linhas(self):
        """
        A soma dos totais tem de ser o tamanho do recorte. Um balde que escapasse da
        classificação sumiria da tela sem deixar rastro — a tabela mostraria menos
        documentos do que existem e ninguém saberia quais faltaram.
        """
        linhas = views._resumo_por_ies(self.df)
        self.assertEqual(sum(linha['total'] for linha in linhas), len(self.df))
        for linha in linhas:
            soma = sum(linha[chave] for chave in views.CHAVE_DO_BALDE.values())
            self.assertEqual(soma, linha['total'], linha['ies'])

    def test_a_classificacao_e_a_mesma_da_vista_de_beneficiarios(self):
        """
        Uma regra só, `_balde_do_documento`, para as duas vistas. Este teste compara o
        resultado agrupado com a contagem crua da MESMA função: se alguém reescrever a
        classificação aqui dentro, os dois números divergem.
        """
        esperado = views._balde_do_documento(self.df).value_counts()
        linhas = views._resumo_por_ies(self.df)
        for balde, chave in views.CHAVE_DO_BALDE.items():
            self.assertEqual(sum(linha[chave] for linha in linhas),
                             int(esperado.get(balde, 0)), balde)

    def test_beneficiario_e_pessoa_e_nao_linha(self):
        """
        A IES A tem três linhas e DUAS pessoas: o CPF 1 mandou contrato e RIAF, e isso é
        uma pessoa com dois documentos, não duas pessoas.
        """
        por_ies = {linha['ies']: linha for linha in views._resumo_por_ies(self.df)}
        self.assertEqual(por_ies['IES A']['beneficiarios'], 2)
        self.assertEqual(por_ies['IES A']['total'], 3)
        self.assertEqual(por_ies['IES B']['beneficiarios'], 1)

    def test_inadimplente_nao_conta_como_beneficiario(self):
        """
        Quem só aparece por inadimplência NÃO é beneficiário. O caso mais gritante é o
        balde `Inadimplentes`: ele nem é documento nosso — é cobrança injetada do
        relatório do site, de semestre em que o aluno não teve lançamento nenhum.
        Contá-lo era dizer que a OVG atende quem ela não custeou.

        Medido na base real (2025-1): 7.541 CPFs saem por esta regra, 7.488 deles vindos
        só desse balde.
        """
        df = _abas(
            faculdade=['IES A', 'IES A'],
            cpf=['1', '2'],
            documento=['CONTRATO', 'CONTRATO'],
            status_ia=['Válido', 'Inadimplente'],
            documento_ausente=['NÃO', 'SIM'],
            veredito_documento=['VÁLIDO', ''],
        )
        linha = views._resumo_por_ies(df)[0]
        self.assertEqual(linha['beneficiarios'], 1)
        # A LINHA continua contando: documento e pessoa são bases diferentes, e a
        # cobrança indevida é justamente o que a coluna `Inadimplentes` denuncia.
        self.assertEqual(linha['total'], 2)
        self.assertEqual(linha['Inadimplentes'], 1)

    def test_os_tres_baldes_de_inadimplencia_saem_juntos_da_contagem(self):
        """
        Não é só o balde injetado: `Inadimplentes Proc.` e `Inadimplentes Não Proc.`
        também são pessoas sem repasse líquido no semestre. Os três descrevem o mesmo
        fato sobre o dinheiro, e o desempate entre eles é sobre a LEITURA do documento.
        """
        df = _abas(
            faculdade=['IES A'] * 3,
            cpf=['1', '2', '3'],
            documento=['CONTRATO', 'RIAF', 'HISTÓRICO'],
            status_ia=['Inadimplente'] * 3,
            documento_ausente=['SIM', 'NÃO', 'NÃO'],
            veredito_documento=['', 'VÁLIDO', ''],
        )
        balde = views._balde_do_documento(df)
        self.assertEqual(set(balde), set(views.BALDES_INADIMPLENTES))
        self.assertEqual(views._resumo_por_ies(df)[0]['beneficiarios'], 0)

    def test_quem_tem_documento_normal_continua_contando(self):
        """
        O CPF MISTO — linha inadimplente e linha normal — é, em regra, a inscrição
        estornada convivendo com a que de fato recebeu (ver
        `test_inadimplente_por_estorno.py`). A pessoa é beneficiária pela segunda;
        apagá-la pela primeira descontaria o mesmo estorno duas vezes.
        """
        df = _abas(
            faculdade=['IES A', 'IES A'],
            cpf=['1', '1'],
            documento=['CONTRATO', 'RIAF'],
            status_ia=['Inadimplente', 'Válido'],
            documento_ausente=['SIM', 'NÃO'],
            veredito_documento=['', 'VÁLIDO'],
        )
        self.assertEqual(views._resumo_por_ies(df)[0]['beneficiarios'], 1)

    def test_o_total_segue_a_mesma_regra_da_coluna(self):
        """
        O chip do topo e a coluna embaixo dele têm de contar a mesma coisa. O total NÃO
        é a soma da coluna (quem muda de IES aparece nas duas), mas é a mesma DEFINIÇÃO
        aplicada ao recorte inteiro.
        """
        df = _abas(
            faculdade=['IES A', 'IES B', 'IES B'],
            cpf=['1', '1', '2'],
            documento=['CONTRATO', 'CONTRATO', 'CONTRATO'],
            status_ia=['Válido', 'Válido', 'Inadimplente'],
            documento_ausente=['NÃO', 'NÃO', 'SIM'],
            veredito_documento=['VÁLIDO', 'VÁLIDO', ''],
        )
        linhas = views._resumo_por_ies(df)
        self.assertEqual(sum(l['beneficiarios'] for l in linhas), 2)  # o CPF 1 nas duas
        atendidos = df[~views._balde_do_documento(df).isin(views.BALDES_INADIMPLENTES)]
        self.assertEqual(int(atendidos['cpf'].nunique()), 1)          # mas é 1 pessoa

    def test_a_ordem_inicial_e_a_maior_primeiro(self):
        """
        ~110 instituições numa tabela: a ordem inicial é o que decide se a primeira tela
        mostra onde está a massa ou uma faculdade qualquer.
        """
        linhas = views._resumo_por_ies(self.df)
        self.assertEqual([linha['ies'] for linha in linhas], ['IES A', 'IES B'])

    def test_recorte_vazio_nao_estoura(self):
        """Filtro que não casa com nada devolve lista vazia, e não exceção."""
        self.assertEqual(views._resumo_por_ies(self.df.iloc[0:0]), [])


class TestFiltroDeDocumento(SimpleTestCase):
    """O recorte por tipo de documento — o filtro que só a vista por IES tem."""

    def setUp(self):
        self.df = _abas(
            faculdade=['IES A', 'IES A', 'IES A'],
            cpf=['1', '1', '2'],
            documento=['CONTRATO', 'RIAF', 'HISTÓRICO'],
            status_ia=['Válido', 'Válido', 'Ausente'],
            documento_ausente=['NÃO', 'NÃO', 'NÃO'],
            veredito_documento=['VÁLIDO', 'VÁLIDO', ''],
        )

    def test_sem_parametro_vale_tudo(self):
        """Estado inicial da tela: nenhum documento marcado é o panorama, não o vazio."""
        self.assertEqual(len(views._aplicar_filtro_de_documentos(self.df, Pedido())), 3)

    def test_recorta_pelos_rotulos_pedidos(self):
        recorte = views._aplicar_filtro_de_documentos(
            self.df, Pedido(documentos='CONTRATO||RIAF'))
        self.assertEqual(sorted(recorte['documento']), ['CONTRATO', 'RIAF'])

    def test_rotulo_desconhecido_nao_devolve_a_base_inteira(self):
        """
        Pedir um documento que não existe tem de dar vazio. Ignorar o rótulo devolveria
        as três linhas para quem pediu uma — o oposto do que se pediu, e silencioso.
        """
        self.assertEqual(len(views._aplicar_filtro_de_documentos(
            self.df, Pedido(documentos='INEXISTENTE'))), 0)

    def test_o_separador_e_barra_dupla(self):
        """
        `||` e não vírgula, igual ao filtro de IES: é o mesmo parâmetro que a tabela de
        beneficiários já usava, e trocar o separador calaria os dois de uma vez.
        """
        self.assertEqual(len(views._aplicar_filtro_de_documentos(
            self.df, Pedido(documentos='CONTRATO,RIAF'))), 0)


class TestApiResumoIES(TestCase):
    """A rota: permissão e formato da resposta."""

    @classmethod
    def setUpTestData(cls):
        cls.com_acesso = Usuario.objects.create_user(
            usuario="com.acesso.ies@ovg.org.br", password="x", nome="Com Acesso",
            p_dash_documentos_ia=True,
        )
        cls.sem_acesso = Usuario.objects.create_user(
            usuario="sem.acesso.ies@ovg.org.br", password="x", nome="Sem Acesso",
            p_dash_documentos_ia=False,
        )

    def setUp(self):
        self.cliente = Client()
        self.url = reverse("dash_documentos_ia_resumo_ies")

    def test_exige_login(self):
        self.assertEqual(self.cliente.get(self.url).status_code, 302)

    def test_recusa_quem_nao_tem_permissao(self):
        self.cliente.force_login(self.sem_acesso)
        self.assertEqual(self.cliente.get(self.url).status_code, 403)

    def test_responde_no_formato_que_a_tela_espera(self):
        """
        As chaves são contrato com o JavaScript: `CHAVES_DAS_FATIAS` no front lê
        exatamente estes nomes, na mesma ordem das seis fatias.
        """
        self.cliente.force_login(self.com_acesso)
        corpo = self.cliente.get(self.url).json()
        self.assertEqual(corpo['status'], 'ok')
        self.assertIsInstance(corpo['linhas'], list)
        for chave in list(views.CHAVE_DO_BALDE.values()) + ['beneficiarios', 'ies']:
            self.assertIn(chave, corpo['totais'])
        for linha in corpo['linhas']:
            self.assertIn('ies', linha)
            self.assertEqual(
                sum(linha[chave] for chave in views.CHAVE_DO_BALDE.values()),
                linha['total'], linha['ies'])


class TestPecasDaVistaNaTela(TestCase):
    """O que o JavaScript procura na página tem de existir."""

    @classmethod
    def setUpTestData(cls):
        cls.usuario = Usuario.objects.create_user(
            usuario="tela.ies@ovg.org.br", password="x", nome="Tela",
            p_dash_documentos_ia=True,
        )

    def setUp(self):
        self.cliente = Client()
        self.cliente.force_login(self.usuario)
        self.html = self.cliente.get(reverse("dash_documentos_ia")).content.decode()

    def test_a_vista_publica_os_elementos_que_o_js_preenche(self):
        for elemento in ['id="ies-chips"', 'id="tabela-ies"', 'id="ies-cabecalho"',
                         'id="ies-corpo"', 'id="ies-contagem"', 'id="ies-busca"',
                         'id="ies-filtros"']:
            self.assertIn(elemento, self.html, elemento)

    def test_o_placeholder_de_em_construcao_saiu(self):
        """A vista existe de verdade; o aviso de que ela viria não pode sobreviver a ela."""
        self.assertNotIn('Em construção', self.html)

    def test_o_filtro_de_documento_tem_os_cinco_e_nasce_recolhido(self):
        """
        Cinco opções, uma por aba do motor. E `display: none` no markup: ele só recorta a
        vista por IES, e a tela abre em Beneficiários — visível ali, aceitaria cliques
        que não mudariam número nenhum.
        """
        import re

        valores = re.findall(
            r'<input type="checkbox" value="([^"]+)" class="filter-documento-ies peer',
            self.html)
        self.assertEqual(valores, ['CONTRATO', 'RIAF', 'HISTÓRICO',
                                   'BENEFÍCIOS', 'FINANCIAMENTO'])
        secao = self.html.split('id="filtro-documentos"', 1)[1].split('>', 1)[0]
        self.assertIn('display: none', secao)

    def test_contrato_nasce_marcado(self):
        """
        A tela abre em CONTRATO, e não em "todos". "Todos" numa linha por instituição é
        a soma de cinco perguntas diferentes na mesma célula — o estado que menos
        responde. O contrato é o de maior cobertura: todos os semestres, 84 mil linhas.
        """
        import re

        marcados = re.findall(
            r'<input type="checkbox" value="([^"]+)" class="filter-documento-ies peer'
            r' sr-only" checked>', self.html)
        self.assertEqual(marcados, ['CONTRATO'])

    def test_a_ordenacao_mora_no_cabecalho_da_tabela_e_so_ali(self):
        """
        Houve uma seção "Ordenar por" na barra, com as perguntas nomeadas (menor % de
        pendência, maior nº de pendências...). Ela saiu: o cabeçalho de cada coluna já
        ordena, e os seis chips do topo também. Três caminhos para o mesmo gesto é um a
        mais do que a tela precisa — e o da barra era o único que exigia abrir a barra.

        A tabela abre em ordem ALFABÉTICA e é para lá que "Restaurar Padrão" volta;
        ordenar é gesto sobre a tabela, não recorte guardado na barra.
        """
        self.assertNotIn('filtro-ordem', self.html)
        self.assertNotIn('filter-ordem-ies', self.html)
        # E o caminho que restou continua de pé: o cabeçalho é escrito pelo JS dentro
        # desta linha, e é nela que o clique de ordenação é delegado.
        self.assertIn('id="ies-cabecalho"', self.html)

    def test_os_cinco_valores_sao_os_rotulos_que_a_api_conhece(self):
        """
        O `value` do checkbox viaja cru no parâmetro `documentos`, e a view compara com
        `ABA_POR_ROTULO`. Um valor fora dessa lista é descartado em silêncio do outro
        lado — a tela recortaria para o vazio sem dizer por quê.
        """
        import re

        valores = re.findall(
            r'<input type="checkbox" value="([^"]+)" class="filter-documento-ies peer',
            self.html)
        self.assertEqual(sorted(valores), sorted(views.ABA_POR_ROTULO))

    def test_periodo_e_instituicao_valem_nos_dois_modos(self):
        """
        Os dois recortam o UNIVERSO, não o sujeito: "2025-1" é a mesma pergunta feita ao
        aluno e à faculdade. Dentro de `#filtros-beneficiarios` eles sumiriam junto com o
        bloco no modo IES, e a vista somaria os quatro semestres de uma vez.
        """
        bloco = self.html.split('id="filtros-beneficiarios"', 1)[1].split('</aside>', 1)[0]
        # O bloco vai até o `</div>` que o fecha; basta que os dois controles não
        # estejam nele antes da seção de Instituição, que vem depois.
        self.assertNotIn('filter-semestre', bloco.split('id="filtro-documentos"', 1)[0])
        self.assertIn('filter-semestre', self.html)
        self.assertIn('openModalIES()', self.html)


class TestSemanticaDoPercentualNaTabela(SimpleTestCase):
    """
    O PERCENTUAL DE CADA COLUNA — a regra vive no JavaScript (`COLUNAS_IES`), e este
    teste lê o arquivo. É verificação de texto, e não de comportamento; existe porque a
    regra é fácil de inverter sem querer e o erro é silencioso: os números continuam
    aparecendo, só que dizendo o contrário do que dizem.

    A REGRA: o percentual diz quanto daquela dimensão já está RESOLVIDO — maior é
    melhor. Zero pendências vale 100% enviado. As três colunas de inadimplência são a
    exceção e mostram a fatia crua, onde maior é PIOR: elas não são um passo do caminho
    que se completa, são um erro em curso, e a meta é levá-las a zero.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        import os

        caminho = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'static', 'dash_documentos_ia', 'js', 'dash_documentos_ia.js')
        with open(caminho, encoding='utf-8') as arquivo:
            cls.js = arquivo.read()

    def test_as_colunas_de_problema_mostram_progresso_e_nao_a_fatia(self):
        """
        `progresso(falta, base)` é `1 - falta/base`: com `falta` em zero ele dá 100%.
        É o que faz "0 pendentes" ler como "enviou tudo" em vez de "0,0%".
        """
        self.assertIn(
            "const progresso = (falta, base) => (base > 0 ? (1 - (falta || 0) / base) * 100 : null);",
            self.js)
        self.assertIn("medida: '% enviado',", self.js)
        self.assertIn("pct: (l) => progresso(l.NaoEnviados, esperadosDe(l)) },", self.js)
        self.assertIn("medida: '% já processado',", self.js)
        self.assertIn("pct: (l) => progresso(l.NaoProcessados, enviadosDe(l)) },", self.js)

    def test_os_tres_de_inadimplencia_sao_a_excecao(self):
        """Fatia crua e marcados com `inverso`, que é o que os pinta como alerta."""
        for chave in ['InadProc', 'InadNaoProc', 'Inadimplentes']:
            with self.subTest(coluna=chave):
                self.assertIn("pct: (l) => fatia(l.%s, l.total) }" % chave, self.js)
        self.assertEqual(self.js.count("medida: '% do total', inverso: true,"), 3)

    def test_a_cobranca_sem_lastro_fica_fora_do_que_a_ies_deve(self):
        """
        `Inadimplentes` é cobrança injetada do relatório do site, de semestre sem
        lançamento nenhum — não é documento que a IES deva. Dentro do denominador, ela
        puniria no `% enviado` justamente quem foi cobrado errado.
        """
        self.assertIn(
            "const esperadosDe = (linha) => (linha.total || 0) - (linha.Inadimplentes || 0);",
            self.js)

    def test_base_zero_nao_inventa_percentual(self):
        """Sem base não há progresso a medir; a célula sai só com o número."""
        self.assertIn("const fatia = (valor, base) => (base > 0 ? ((valor || 0) / base) * 100 : null);",
                      self.js)
        self.assertIn("const pct = valorPct === null ? '' :", self.js)

    def test_o_cabecalho_nomeia_a_medida_de_cada_coluna(self):
        """
        Sem o rótulo, "36 (97,2%)" sob PENDENTES lê como "97,2% estão pendentes" — o
        oposto do que diz. Cada coluna tem denominador próprio, e é o cabeçalho que o
        informa.
        """
        self.assertIn('docia-ies-th__medida', self.js)
        for medida in ["'% do esperado'", "'% já processado'", "'% enviado'", "'% do total'"]:
            with self.subTest(medida=medida):
                self.assertIn('medida: %s' % medida, self.js)


class TestExportacaoDaVisaoIES(TestCase):
    """
    O botão Excel da tabela por instituição.

    POR QUÊ EXISTE A ROTA: aqui não há teto a resolver como no card de beneficiários —
    são ~110 linhas e todas cabem na tela. A planilha é a resposta para LEVAR: cruzar
    com outra base, mandar por e-mail e guardar o recorte de um período que a próxima
    execução do motor vai sobrescrever.

    O QUE ESTE TESTE GARANTE: que o arquivo é o que está na tela. Se a rota montasse a
    própria cadeia de filtros, bastaria um filtro novo entrar num lado e não no outro
    para o download deixar de ser o recorte que a pessoa estava vendo — sem erro em
    lugar nenhum, e só descoberto por quem conferisse número a número.
    """

    @classmethod
    def setUpTestData(cls):
        cls.com_acesso = Usuario.objects.create_user(
            usuario="exporta.ies@ovg.org.br", password="x", nome="Exporta",
            p_dash_documentos_ia=True,
        )
        cls.sem_acesso = Usuario.objects.create_user(
            usuario="sem.exporta.ies@ovg.org.br", password="x", nome="Sem",
            p_dash_documentos_ia=False,
        )

    def setUp(self):
        self.cliente = Client()
        self.url = reverse("dash_documentos_ia_exportar_ies")

    def test_exige_login(self):
        self.assertEqual(self.cliente.get(self.url).status_code, 302)

    def test_recusa_quem_nao_tem_permissao(self):
        self.cliente.force_login(self.sem_acesso)
        self.assertEqual(self.cliente.get(self.url).status_code, 403)

    def test_devolve_uma_planilha_com_nome_legivel(self):
        """
        O nome acompanha o título do card na tela. O de máquina
        (`envios-ies-2026-09-02-1121`) obriga quem recebe por e-mail a abrir o arquivo
        para saber o que é.
        """
        self.cliente.force_login(self.com_acesso)
        resposta = self.cliente.get(self.url)
        self.assertEqual(resposta.status_code, 200)
        self.assertIn('spreadsheetml', resposta['Content-Type'])
        self.assertIn('Envios e Pendencias por Instituicao',
                      resposta['Content-Disposition'])

    def test_o_arquivo_traz_as_mesmas_linhas_que_a_api_da_tela(self):
        """
        A rota repete a cadeia de `api_resumo_ies`. Divergindo, o arquivo baixado
        deixaria de ser o recorte que está na tela.
        """
        import io as _io

        import openpyxl

        self.cliente.force_login(self.com_acesso)
        consulta = '?semestres=2025-1&documentos=CONTRATO'
        planilha = openpyxl.load_workbook(
            _io.BytesIO(self.cliente.get(self.url + consulta).content)).active
        da_tela = self.cliente.get(
            reverse("dash_documentos_ia_resumo_ies") + consulta).json()['linhas']

        # -1 pelo cabeçalho.
        self.assertEqual(planilha.max_row - 1, len(da_tela))
        nomes_no_arquivo = {linha[0] for linha in
                            planilha.iter_rows(min_row=2, max_col=1, values_only=True)}
        self.assertEqual(nomes_no_arquivo, {linha['ies'] for linha in da_tela})

    def test_leva_as_duas_bases_dos_percentuais(self):
        """
        `Esperados` e `Enviados` não são colunas da tela — lá os percentuais já saem
        calculados ao lado de cada número. No arquivo elas vão junto porque, sem elas,
        quem for conferir teria de reconstruir a regra de cabeça: `Esperados` tira a
        cobrança sem lastro do que a IES realmente deve, e é o passo que não se adivinha
        olhando as outras colunas.
        """
        import io as _io

        import openpyxl

        self.cliente.force_login(self.com_acesso)
        planilha = openpyxl.load_workbook(
            _io.BytesIO(self.cliente.get(self.url).content)).active
        cabecalho = [c.value for c in planilha[1]]
        self.assertEqual(cabecalho[-3:], ['Total de Documentos', 'Esperados', 'Enviados'])

        for linha in planilha.iter_rows(min_row=2, values_only=True):
            (_, _, proc, nao_proc, pend, inad_p, inad_np, inad,
             total, esperados, enviados) = linha
            with self.subTest(ies=linha[0]):
                self.assertEqual(proc + nao_proc + pend + inad_p + inad_np + inad, total)
                self.assertEqual(esperados, total - inad)
                self.assertEqual(enviados, esperados - pend)
